from __future__ import annotations

import argparse
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-url", required=True)
    parser.add_argument("--mvv", default=str(Path(__file__).resolve().parents[2] / "input" / "MVV.xlsx"))
    parser.add_argument("--rd", default=str(Path(__file__).resolve().parents[2] / "input" / "RD.txt"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    download_path = Path(tempfile.gettempdir()) / "browser-smoke.xlsx"
    mvv_only_download_path = Path(tempfile.gettempdir()) / "browser-smoke-mvv-only.xlsx"
    rd_only_download_path = Path(tempfile.gettempdir()) / "browser-smoke-rd-only.xlsx"
    rd_only_input_path = Path(tempfile.gettempdir()) / "browser-smoke-rd-only.txt"
    rd_only_input_path.write_text("L_1,,10,20,292\nE-1,,11,21,292\nL_2,,12,22,280\n", encoding="utf-8")

    for path in [download_path, mvv_only_download_path, rd_only_download_path]:
        if path.exists():
            path.unlink()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(accept_downloads=True)
        page.goto(args.base_url, wait_until="domcontentloaded")
        page.wait_for_selector("#statusText")
        page.wait_for_function("document.querySelector('.brand-logo')?.getAttribute('src') === './assets/openblast-logo.png'")
        page.wait_for_function("document.documentElement.lang === 'pt-BR'")
        page.wait_for_function("document.querySelector('#languageSelect')?.value === 'pt'")
        page.wait_for_function("document.querySelector('#appSubtitle')?.hidden === true")
        page.wait_for_function("document.querySelector('#workflowStep1')?.textContent === 'Anexar Planejado'")
        page.wait_for_function("document.querySelector('#workflowStep2')?.textContent === 'Anexar Realizado'")
        page.wait_for_function("document.querySelector('#mvvFileLabel')?.textContent === 'PLANEJADO.xlsx'")
        page.wait_for_function("document.querySelector('#rdFileLabel')?.textContent === 'REALIZADO.txt'")
        page.wait_for_function("document.querySelector('#mvvFileHint')?.textContent === 'Plano de Perfuração'")
        page.wait_for_function("document.querySelector('#rdFileHint')?.textContent === 'Arquivo de Coordenadas da Topografia'")
        page.wait_for_function("document.querySelector('#statusText')?.textContent === 'Anexe PLANEJADO.xlsx para organizar PLANEJADO ou anexe tambem REALIZADO.txt para consolidar.'")
        page.wait_for_function("document.querySelector('#generateBtn')?.textContent === 'Gerar Dado Consolidado Planejado vs Realizado'")
        page.wait_for_function("document.querySelector('#rdOnlyBtn')?.textContent === 'Organize Somente o Executado'")
        page.wait_for_function("document.querySelector('#mvvOnlyBtn')?.textContent === 'Organize Somente o Dado Planejado'")

        page.locator("#languageSelect").select_option("en")
        page.wait_for_function("document.documentElement.lang === 'en'")
        page.wait_for_function("document.querySelector('#languageSelect')?.value === 'en'")
        page.wait_for_function("document.querySelector('#generateBtn')?.textContent === 'Generate workbook'")
        page.wait_for_function("document.querySelector('#statusText')?.textContent?.includes('Attach MVV.xlsx to organize MVV')")

        page.locator("#languageSelect").select_option("zh")
        page.wait_for_function("document.documentElement.lang === 'zh-Hans'")
        page.wait_for_function("document.querySelector('#languageSelect')?.value === 'zh'")
        page.wait_for_function("document.querySelector('#generateBtn')?.textContent === '生成工作簿'")

        page.locator("#languageSelect").select_option("pt")
        page.wait_for_function("document.documentElement.lang === 'pt-BR'")
        page.wait_for_function("document.querySelector('#languageSelect')?.value === 'pt'")
        page.wait_for_function("document.querySelector('#generateBtn')?.textContent === 'Gerar Dado Consolidado Planejado vs Realizado'")

        page.set_input_files("#mvvFile", args.mvv)
        page.locator("#mvvOnlyBtn").wait_for(state="visible")
        page.wait_for_function("!document.querySelector('#mvvOnlyBtn')?.disabled")
        page.get_by_role("button", name="Organize Somente o Dado Planejado").click()
        page.wait_for_function("document.querySelector('#statusText')?.textContent?.includes('Plano MVV organizado.')")
        page.locator("#downloadLink").wait_for(state="visible")
        with page.expect_download() as mvv_only_download_info:
            page.get_by_role("link", name="Baixar MVV_PLANO_PERFURACAO_ORGANIZADO.xlsx").click()
        mvv_only_download = mvv_only_download_info.value
        mvv_only_download.save_as(str(mvv_only_download_path))

        page.set_input_files("#rdFile", args.rd)
        page.locator("#generateBtn").wait_for(state="visible")
        page.wait_for_function("!document.querySelector('#generateBtn')?.disabled")
        page.get_by_role("button", name="Gerar Dado Consolidado Planejado vs Realizado").click()
        page.wait_for_function("document.querySelector('#statusText')?.textContent?.includes('Planilha gerada.')")
        page.locator("#downloadLink").wait_for(state="visible")
        with page.expect_download() as download_info:
            page.get_by_role("link", name="Baixar MVV_RD_CONSOLIDADO_FINAL.xlsx").click()
        download = download_info.value
        download.save_as(str(download_path))

        page.set_input_files("#rdFile", str(rd_only_input_path))
        page.wait_for_function("!document.querySelector('#rdOnlyBtn')?.disabled")
        page.once("dialog", lambda dialog: dialog.accept("10"))
        page.get_by_role("button", name="Organize Somente o Executado").click()
        page.wait_for_function("document.querySelector('#statusText')?.textContent?.includes('Executado organizado.')")
        page.locator("#downloadLink").wait_for(state="visible")
        with page.expect_download() as rd_only_download_info:
            page.get_by_role("link", name="Baixar RD_EXECUTADO_ORGANIZADO.xlsx").click()
        rd_only_download = rd_only_download_info.value
        rd_only_download.save_as(str(rd_only_download_path))
        browser.close()

    wb = load_workbook(download_path, data_only=True)
    assert wb.sheetnames == ["CONSOLIDADO_FINAL", "RD_TRATADA", "LOG_VALIDACAO"]
    download_path.unlink(missing_ok=True)

    mvv_only_wb = load_workbook(mvv_only_download_path, data_only=True)
    assert mvv_only_wb.sheetnames == ["PLANO_MVV"]
    headers = [cell.value for cell in mvv_only_wb["PLANO_MVV"][1]]
    assert headers == ["ID", "Type", "Explosivo", "Diameter", "X Collar", "Y Collar", "Z Collar", "Depth", "Sub Drill", "Azimuth", "Dip", "Tampao", "Carga"]
    mvv_only_download_path.unlink(missing_ok=True)

    rd_only_wb = load_workbook(rd_only_download_path, data_only=True)
    assert rd_only_wb.sheetnames == ["RD_EXECUTADO"]
    rd_headers = [cell.value for cell in rd_only_wb["RD_EXECUTADO"][1]]
    assert rd_headers == ["ID", "Y", "X", "Z", "Profundidade"]
    assert rd_only_wb["RD_EXECUTADO"][2][0].value == 1
    assert rd_only_wb["RD_EXECUTADO"][2][1].value == 11
    assert rd_only_wb["RD_EXECUTADO"][2][2].value == 21
    assert rd_only_wb["RD_EXECUTADO"][2][3].value == 292
    assert rd_only_wb["RD_EXECUTADO"][2][4].value == 10
    assert rd_only_wb["RD_EXECUTADO"][3][0].value == 2
    assert rd_only_wb["RD_EXECUTADO"][3][4].value == 10
    rd_only_download_path.unlink(missing_ok=True)
    rd_only_input_path.unlink(missing_ok=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
