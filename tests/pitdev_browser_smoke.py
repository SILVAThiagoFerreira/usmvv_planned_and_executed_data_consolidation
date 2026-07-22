from __future__ import annotations

import argparse
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-url", required=True)
    parser.add_argument("--field", required=True)
    parser.add_argument("--plan", required=True)
    args = parser.parse_args()
    download_path = Path(tempfile.gettempdir()) / "pitdev-browser-smoke.xlsx"
    screenshot_path = Path(tempfile.gettempdir()) / "pitdev-browser-smoke.png"
    download_path.unlink(missing_ok=True)
    screenshot_path.unlink(missing_ok=True)

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page(accept_downloads=True, viewport={"width": 1440, "height": 1100})
        page.goto(args.base_url, wait_until="domcontentloaded")
        page.on("console", lambda message: print(f"browser console: {message.type}: {message.text}"))
        page.wait_for_selector("#pitdevTitle")
        page.wait_for_function("document.querySelector('#pitdevTitle')?.textContent === 'Consolidação de Projeto para O-PìtDev'")
        page.wait_for_function("document.querySelector('#pitdevGenerateBtn')?.disabled === true")
        page.set_input_files("#pitdevFieldFile", args.field)
        page.set_input_files("#pitdevPlanFile", args.plan)
        page.wait_for_function("!document.querySelector('#pitdevGenerateBtn')?.disabled")
        page.wait_for_function("document.querySelector('#pitdevStatusText')?.textContent?.includes('Pronto para consolidar')")
        page.get_by_role("button", name="Consolidar Projeto para O-PìtDev").click()
        page.wait_for_timeout(2000)
        print("pitdev status:", page.locator("#pitdevStatusText").text_content())
        print("pitdev log:", page.locator("#pitdevLogOutput").text_content())
        page.wait_for_function("document.querySelector('#pitdevStatusText')?.textContent?.includes('Consolidação gerada.')")
        page.wait_for_function("document.querySelector('#pitdevSummaryCards')?.textContent?.includes('24')")
        page.locator("#pitdevDownloadLink").wait_for(state="visible")
        page.screenshot(path=str(screenshot_path), full_page=True)
        with page.expect_download() as download_info:
            page.get_by_role("link", name="Baixar CONSOLIDACAO_PROJETO_O-PITDEV.xlsx").click()
        download_info.value.save_as(str(download_path))
        browser.close()

    workbook = load_workbook(download_path, data_only=True)
    assert workbook.sheetnames == ["CONSOLIDACAO_O-PITDEV", "LOG_O-PITDEV"]
    sheet = workbook["CONSOLIDACAO_O-PITDEV"]
    assert [cell.value for cell in sheet[1]] == [
        "ID", "Y", "X", "Z", "Diâmetro", "Azimute", "Ângulo planejado", "Ângulo do talude"
    ]
    assert sheet.max_row == 25
    assert sheet[2][0].value == 97
    assert sheet[2][1].value == 8929912.804
    assert sheet[2][4].value == 5
    assert sheet[2][6].value == 0
    assert sheet[2][7].value == 90
    print(f"ok - O-PitDev browser flow | rows={sheet.max_row - 1} | screenshot={screenshot_path}")
    download_path.unlink(missing_ok=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
